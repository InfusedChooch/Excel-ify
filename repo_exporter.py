#!/usr/bin/env python3
# /tools/repo_exporter.py

from __future__ import annotations

import mimetypes
import os
import sys
import textwrap
from datetime import datetime
from pathlib import Path
from typing import Counter, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

# ─── Configuration ────────────────────────────────────────────────────────
EXCLUDED_DIRS: set[str] = {
    "venv",
    ".venv",
    ".git",
    "__pycache__",
    ".idea",
    ".vscode",
    "node_modules",
}

INCLUDE_EXTS = {
    ".py",
    ".md",
    ".txt",
    ".json",
    ".yaml",
    ".yml",
    ".toml",
    ".ini",
    ".cfg",
    ".csv",
    ".tsv",
    ".js",
    ".ts",
    ".html",
    ".css",
    ".sh",
    ".bat",
    ".ps1",
    ".sql",
}

README_CANDIDATES = {"readme", "read_me"}

# try to enable black auto-format (idea #7)
try:
    import black

    def fmt_py(src: str) -> str:  # noqa: D401
        try:
            return black.format_str(src, mode=black.FileMode())
        except black.NothingChanged:
            return src

except ImportError:

    def fmt_py(src: str) -> str:  # noqa: D401
        return src  # silently no-op if black not available


# ─── Helpers ──────────────────────────────────────────────────────────────
def is_binary(path: Path) -> bool:
    """Heuristic: MIME + NUL sniff."""
    mime, _ = mimetypes.guess_type(str(path))
    if mime and not mime.startswith("text/"):
        return True
    try:
        with path.open("rb") as f:
            return b"\x00" in f.read(512)
    except Exception:
        return True


def iter_files(root: Path):
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in EXCLUDED_DIRS]
        for name in filenames:
            p = Path(dirpath) / name
            if p.suffix.lower() in INCLUDE_EXTS and not is_binary(p):
                yield p.relative_to(root)


def sheet_safe(name: str) -> str:
    """Excel sheet names max 31 chars, no /:*?[]"""
    return (
        name.replace(os.sep, "_").replace(":", "_")[:31].encode("ascii", "replace").decode()
    )


def top_level(path: Path) -> str:
    return path.parts[0] if len(path.parts) > 1 else "_root"


# ─── Main ─────────────────────────────────────────────────────────────────
def build_excel(root: Path, out_file: Path):
    files = list(iter_files(root))
    if not files:
        print("❌ No matching files found.")
        return

    wb = Workbook()
    wb.remove(wb.active)  # start fresh

    # ── Summary sheet ────────────────────────────────────────────────────
    summary = wb.create_sheet("Summary")
    summary.append(["Relative Path", "Language", "Size (bytes)", "Last Modified", "Sheet"])
    for c in summary[1]:
        c.font = Font(bold=True)

    lang_map: Dict[str, str] = {
        ".py": "Python",
        ".js": "JavaScript",
        ".ts": "TypeScript",
        ".md": "Markdown",
        ".json": "JSON",
        ".yaml": "YAML",
        ".yml": "YAML",
        ".html": "HTML",
        ".css": "CSS",
        ".sh": "Shell",
        ".bat": "Batch",
        ".ps1": "PowerShell",
        ".sql": "SQL",
        ".csv": "CSV",
        ".tsv": "TSV",
        ".toml": "TOML",
        ".ini": "INI",
        ".cfg": "Config",
        ".txt": "Text",
    }

    # keep quick stats (idea #5)
    total_loc = 0
    by_lang: Counter[str] = Counter()

    # one sheet per *top-level* directory (idea #3)
    dir_sheets: Dict[str, any] = {}
    fenced_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    mono_font = Font(name="Consolas")
    header_font = Font(bold=True)

    for p in files:
        abs_p = root / p
        ext = p.suffix.lower()
        lang = lang_map.get(ext, ext.lstrip(".").upper())
        by_lang[lang] += 1

        # read content (+black fmt if py & available)
        try:
            src = abs_p.read_text(encoding="utf-8", errors="replace")
            if ext == ".py":
                src = fmt_py(src)
        except Exception as e:
            src = f"<<Error reading file: {e}>>"

        loc = src.count("\n") + 1
        total_loc += loc

        sheet_title = top_level(p)
        if sheet_title not in dir_sheets:
            sh = wb.create_sheet(sheet_safe(sheet_title))
            sh.column_dimensions["A"].width = 8   # line #
            sh.column_dimensions["B"].width = 120  # code
            dir_sheets[sheet_title] = sh
            current_row = 1
        else:
            sh = dir_sheets[sheet_title]
            current_row = sh.max_row + 2  # blank separator

        # section header inside dir sheet
        sh.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        cell = sh.cell(row=current_row, column=1, value=str(p))
        cell.font = header_font
        current_row += 1

        # write code fence + content
        sh.cell(row=current_row, column=2, value=f"```{ext.lstrip('.')}")
        current_row += 1

        trimmed = textwrap.dedent(src.strip("\n"))  # idea #2 (trim leading/trailing blanks)
        for idx, line in enumerate(trimmed.splitlines(), start=1):
            sh.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal="right")
            code_cell = sh.cell(row=current_row, column=2, value=line)
            code_cell.font = mono_font
            code_cell.fill = fenced_fill
            sh.row_dimensions[current_row].height = 14  # idea #2 row height
            current_row += 1

        sh.cell(row=current_row, column=2, value="```")

        # add Summary row (with hyperlink)
        mtime = datetime.fromtimestamp(abs_p.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
        row = [
            str(p),
            lang,
            abs_p.stat().st_size,
            mtime,
            sheet_safe(sheet_title),
        ]
        summary.append(row)
        # hyperlink in first cell (idea #8)
        summary.cell(row=summary.max_row, column=1).hyperlink = abs_p.as_uri()

    # format Summary table (idea #1)
    last_col = get_column_letter(summary.max_column)
    tab = Table(
        displayName="FileSummary",
        ref=f"A1:{last_col}{summary.max_row}",
    )
    style = TableStyleInfo(
        name="TableStyleLight9",
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    summary.add_table(tab)
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = tab.ref

    # auto-sizes
    for idx, col in enumerate(summary.columns, 1):
        max_len = max(len(str(c.value or "")) for c in col)
        summary.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 40)

    # ── Stats sheet (idea #5) ─────────────────────────────────────────────
    stats = wb.create_sheet("Stats")
    stats.append(["Metric", "Value"])
    stats["A2"] = "Total text files"
    stats["B2"] = len(files)
    stats["A3"] = "Total lines of code"
    stats["B3"] = total_loc
    stats["A4"] = "Last modified"
    stats["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    # language breakdown
    stats_row = 6
    stats["A5"] = "Files by language"
    for lang, count in by_lang.most_common():
        stats.cell(row=stats_row, column=1, value=lang)
        stats.cell(row=stats_row, column=2, value=count)
        stats_row += 1
    for c in stats[1]:
        c.font = Font(bold=True)
    stats.column_dimensions["A"].width = 25
    stats.column_dimensions["B"].width = 15

    # ── README sheet (unchanged except new note) ──────────────────────────
    readme = wb.create_sheet("README")
    readme.column_dimensions["A"].width = 110
    readme["A1"] = "📦 Project Source Export"
    readme["A3"] = "⚙️ How this workbook was generated:"
    readme["A4"] = (
        "1. Place repo_exporter.py in the project root\n"
        "2. Run:  python repo_exporter.py   (optionally -o output.xlsx)\n"
        "3. See the Summary sheet for quick navigation (header row is frozen and filterable).\n"
        "4. Each top-level folder has its own sheet with file sections; collapse rows in Excel’s"
        " outline to hide or show individual files.\n"
        "5. If you have the `black` package installed, Python files are auto-formatted."
    )

    # embed dependency files if present
    for req_name in ("requirements.txt", "pyproject.toml", "environment.yml"):
        req_path = root / req_name
        if req_path.exists():
            readme_row = readme.max_row + 2
            readme[f"A{readme_row}"] = f"📃 {req_name}:"
            for i, line in enumerate(req_path.read_text().splitlines(), start=readme_row + 1):
                readme[f"A{i}"] = line.rstrip()

    # ── Save ─────────────────────────────────────────────────────────────
    wb.save(out_file)
    print(f"✅ Export complete → {out_file}")


# ─── CLI ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Export all code / text files in a project to a single Excel workbook."
    )
    parser.add_argument("root", nargs="?", default=".", help="Project root directory")
    parser.add_argument(
        "-o", "--out", default="project_source_export.xlsx", help="Output XLSX filename"
    )
    args = parser.parse_args()

    build_excel(Path(args.root).resolve(), Path(args.out).resolve())
